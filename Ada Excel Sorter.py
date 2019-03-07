# -*- coding: utf-8 -*-
"""
Created on Fri Mar  1 11:45:06 2019

@author: nashfam
"""

import openpyxl
import os
import lxml
import statistics
from openpyxl import Workbook 

wb2 = Workbook()
wb = openpyxl.load_workbook('Sheet.xlsx')
ws1 = wb.active
wsq1 = wb2.create_sheet("Q1 Filtered Sheet")
wsq2 = wb2.create_sheet("Q2 Filtered Sheet")
wsq3 = wb2.create_sheet("Q3 Filtered Sheet")
wsq4 = wb2.create_sheet("Q4 Filtered Sheet")
wsq5 = wb2.create_sheet("Q5 Filtered Sheet")



class CellSorter:
    def __init__(self, ws, ws2,mincol,maxcol, *args):
        self.arguement = args 
        self.ws = ws
        self.ws2 = ws2
        self.col1 = mincol
        self.col2 = maxcol
#Filters out any rows that do not contain the arguments set by the class within 
#the worksheet assigned in ws class arg
    def RowFilter(self):
       row_list = []
       c_row = (0)
       for row in self.ws.iter_rows(min_col=self.col1, max_col=self.col2, max_row=946, min_row=1):
           row_data = []
           c_row += 1
           bad_row = False
           for cell in row:
               row_data.append(cell.value)
           for arg in self.arguement:
               argument = str(arg)
               if len(argument) > 1:
                   c_row_data = str(row_data)
                   if argument not in c_row_data:
                       bad_row = True

                          
           if bad_row == False:
               row_list.append(c_row)
       return row_list
            
#Creates a list of cell position IDs from the output of RowFilter
    def CellAssign(self, columns):
        columns = columns
        columns = columns.upper()
        cell_list = []
        row_list = self.RowFilter()
        for num in row_list:
            row_num = str(num)
            for char in columns:
                cell = char + row_num
                cell_list.append(cell)
        return cell_list
    
#Populates the worksheet accessed by class arg ws2 with the selected cells from CellAssign
#and the columns in the argument
    def FilteredSheet(self, columns, row_start):
        cell_list = self.CellAssign(columns)
        columns = columns
        columns = columns.upper()
        row = row_start
        i_counter = (0)
        counter = (0)
        for string in cell_list:
            char = columns[i_counter]
            cell = self.ws[string]
            cell_data = cell.value
            row_num = str(row)
            filtered_cell = char + row_num
            self.ws2[filtered_cell] = cell_data
            i_counter += 1
            counter += 1
            if counter == len(columns):
                row += 1
                i_counter = (0)
                counter = (0)
        return(row)


#Exclude method that takes 1 or more args and deletes the row if 1 or more of those args are present
    def Exclude(self, *args):
       exclude_list = []
       c_row = (0)
       fil = args
       for row in self.ws2.iter_rows(min_col=self.col1, max_col=self.col2, max_row=946, min_row=0):
           row_data = []
           c_row += 1
           bad_row = False
           for cell in row:
               row_data.append(cell.value)
           for arg in fil:
                if arg in row_data:
                    bad_row = True
           if bad_row == True:
               print(c_row)
               exclude_list.append(c_row)
       for row in reversed(exclude_list):
          self.ws2.delete_rows(row,1)
       return exclude_list

#old version of Exclude but is useful for getting rid of rows with the arg in a sepcific column
    def ColumnExclude(self, col, *args):
        fil = args
        exclude_list = []
        c_row = (0)
        col = col
        for arg in fil:
            c_row = (0)
            for row in self.ws2.iter_rows(min_col=col, max_col=col, max_row=946, min_row=1):
                for cell in row:
                    c_row +=1
                    if cell.value == arg:
                        exclude_list.append(c_row)
                       
        for row in reversed(exclude_list):
            self.ws2.delete_rows(row,1)

        return exclude_list
    
    def RemoveNonstates(self):
        c_row = (0)
        removed_rows = []
        state_names = ['Alabama', 'Alaska', 'Arizona', 'Arkansas', 'California', 'Colorado', 'Connecticut', 'Delaware', 'Florida', 'Georgia', 'Hawaii', 'Idaho', 'Illinois', 'Indiana', 'Iowa', 'Kansas', 'Kentucky', 'Louisiana', 'Maine', 'Maryland', 'Massachusetts', 'Michigan', 'Minnesota', 'Mississippi', 'Missouri', 'Montana', 'Nebraska', 'Nevada', 'New Hampshire', 'New Jersey', 'New Mexico', 'New York', 'North Carolina', 'North Dakota', 'Ohio', 'Oklahoma', 'Oregon', 'Pennsylvania', 'Rhode Island', 'South Carolina', 'South Dakota', 'Tennessee', 'Texas', 'Utah', 'Vermont', 'Virginia', 'Washington', 'West Virginia', 'Wisconsin', 'Wyoming']
        self.ColumnExclude(3,'Lake States')
        for row in self.ws2.iter_rows(min_col=3, max_col=3, max_row=946,min_row=1):
            for cell in row:
                c_row += 1
                bad_row = False
                if cell.value is not None:
                    if cell.value not in state_names:
                        removed_rows.append(c_row)
                        bad_row = True
            if bad_row is True:
                self.ws2.delete_rows(c_row)

        return removed_rows
    
#finds biggest number in a column and returns it                    
    def FindBiggest(self, col):
        col = col
        c_biggest = (0)
        for row in self.ws2.iter_rows(min_col=col, max_col=col, max_row=946, min_row=1):
            for cell in row:
                if cell.value is not None:
                    if cell.value > c_biggest:
                        c_biggest = (cell.value)
        return(c_biggest)
            
#finds the average of all numbers in column                            
    def FindAverage(self, col):
        col = col
        cell_list = []
        for row in self.ws2.iter_rows(min_col=col, max_col=col, max_row=946, min_row=1):
            for cell in row:
                if cell.value is not None:
                    cell_list.append(cell.value)
        return sum(cell_list) / len(cell_list)
    
#finds median number in a column    
    def FindMedian(self, col):
        col = col
        cell_list = []
        for row in self.ws2.iter_rows(min_col=col, max_col=col, max_row=946, min_row=1):
            for cell in row:
                if cell.value is not None:
                    cell_list.append(cell.value)
        return statistics.median(cell_list)
        
                    



            
                
        
        
        
   


# Question 1 filter
q1columns = ('bcdl')
q1filter1 = CellSorter(ws1,wsq1,2,4,2012)
q1filter1.FilteredSheet(q1columns,1)
q1filter1.RemoveNonstates()
print(q1filter1.FindBiggest(12))

# Question 2 filter
q2columns = ('bcdh')
q2filter1 = CellSorter(ws1,wsq2,2,4,1945,'Mountain')
q2filter1.FilteredSheet(q2columns, 1)
q2filter1.RemoveNonstates()

## Question 3 filter
q3columns = ('bcds')
q3filter = CellSorter(ws1,wsq3,2,4)
q3filter.FilteredSheet(q3columns, 1)
exclude_list = []
c_row = (0)
years = list(range(1950,2000))
for row in wsq3.iter_rows(min_col=4, max_col=4, max_row=946, min_row=1):
    for cell in row:
         c_row +=1
         if cell.value not in years:
             exclude_list.append(c_row) 
              
for row in reversed(exclude_list):
     wsq3.delete_rows(row,1)
q3filter.RemoveNonstates()
q3_states =['Nevada', 'Nebraska', 'New Hampshire', 'New Jersey', 'New Mexico', 'New York', 'North Carolina', 'North Dakota']
exclude_list = []
c_row = (0)
for row in wsq3.iter_rows(min_col=3, max_col=3, max_row=946, min_row=1):
    for cell in row:
         c_row +=1
         if cell.value not in q3_states:
             exclude_list.append(c_row)
                
for row in reversed(exclude_list):#delete rows from the bottom up!!
     wsq3.delete_rows(row,1)

## Question 4 filter
q4columns = ('bcdj')
q4filter1 = CellSorter(ws1,wsq4,2,4,1969, 'Northern Plains')
q4filter1_row = q4filter1.FilteredSheet(q4columns, 1)
q4filter2 = CellSorter(ws1,wsq4,2,4,1969, 'Northeast')
q4filter2_row = q4filter2.FilteredSheet(q4columns, q4filter1_row)
q4filter3 = CellSorter(ws1,wsq4,2,4,1969, 'Appalachian')
q4filter3.FilteredSheet(q4columns,q4filter2_row)
q4filter3.RemoveNonstates()
print(q4filter1.FindMedian(10))

## Question 5 filter
q5columns = ('bcdn')
q5filter1 = CellSorter(ws1,wsq5,2,4,2007)
q5filter1.FilteredSheet(q5columns,1)
q5filter1.RemoveNonstates()
c_row = (0)
removed_rows = []
state_names = ['Massachusetts','New Hampshire','Minnesota','North Dakota','Iowa','Maryland','Colorado','Utah', 'Vermont','Washington']

for row in wsq5.iter_rows(min_col=3, max_col=3, max_row=946,min_row=1):
    for cell in row:
        c_row += 1
        bad_row = False
        if cell.value is not None:
            if cell.value not in state_names:
                removed_rows.append(c_row)
              
for num in reversed(removed_rows):
    wsq5.delete_rows(num,1)
    
print(q5filter1.FindAverage(14))
      
wb2.save('Filtered Sheets.xlsx')        
